import requests
import os
import time
import json
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

ANTHROPIC_KEY = os.environ["ANTHROPIC_KEY"]
BASE_URL = "https://recherche-entreprises.api.gouv.fr/search"

# TEST : 10 premières entreprises
ENTREPRISES = [
    "3DEUS DYNAMICS", "A-NSE", "AAS INDUSTRIES", "ABC", "AC DISMANTLING",
    "ACB", "ADB", "ADDEV MATERIALS AEROSPACE SAS", "AddUp", "ADHETEC",
]

def chercher_siren(nom):
    for tentative in range(3):
        try:
            r = requests.get(BASE_URL, params={"q": nom, "per_page": 3}, timeout=15)
            if r.status_code == 429:
                print(f"  Rate limit, attente 60s...")
                time.sleep(60)
                continue
            if r.status_code == 200:
                data = r.json()
                resultats = data.get("results", [])
                if resultats:
                    e = resultats[0]
                    return {
                        "siren": e.get("siren", ""),
                        "nom_trouve": e.get("nom_complet", "") or e.get("nom_raison_sociale", ""),
                        "ville": e.get("siege", {}).get("commune", "") if e.get("siege") else "",
                        "statut": "Trouve"
                    }
                return {"siren": "", "nom_trouve": "", "ville": "", "statut": "Non trouve"}
            return {"siren": "", "nom_trouve": "", "ville": "", "statut": f"Erreur {r.status_code}"}
        except Exception as e:
            return {"siren": "", "nom_trouve": "", "ville": "", "statut": f"Erreur: {str(e)[:40]}"}
    return {"siren": "", "nom_trouve": "", "ville": "", "statut": "Echec apres 3 tentatives"}

def appel_claude_avec_retry(payload, timeout=30):
    for tentative in range(3):
        try:
            r = requests.post(
                "https://api.anthropic.com/v1/messages",
                headers={
                    "x-api-key": ANTHROPIC_KEY,
                    "anthropic-version": "2023-06-01",
                    "content-type": "application/json",
                },
                json=payload,
                timeout=timeout,
            )
            if r.status_code == 429:
                print(f"  Anthropic rate limit, attente 30s... (tentative {tentative+1})")
                time.sleep(30)
                continue
            r.raise_for_status()
            return r.json()
        except Exception as e:
            if tentative == 2:
                raise
            time.sleep(5)
    return None

def verifier_faux_positif(nom_recherche, nom_trouve):
    """Appel Claude SANS web search — juste vérifier si c'est la bonne entreprise."""
    prompt = f"""Est-ce que ces deux noms désignent la même entreprise ?
Nom recherché : "{nom_recherche}"
Nom trouvé : "{nom_trouve}"

Réponds UNIQUEMENT en JSON :
{{"faux_positif": true ou false, "raison": "explication en 1 phrase"}}"""

    payload = {
        "model": "claude-haiku-4-5-20251001",
        "max_tokens": 200,
        "messages": [{"role": "user", "content": prompt}],
    }
    for tentative in range(3):
        try:
            r = requests.post(
                "https://api.anthropic.com/v1/messages",
                headers={"x-api-key": ANTHROPIC_KEY, "anthropic-version": "2023-06-01", "content-type": "application/json"},
                json=payload, timeout=30,
            )
            if r.status_code == 429:
                print(f"  Anthropic 429, attente 30s... (tentative {tentative+1})")
                time.sleep(30)
                continue
            r.raise_for_status()
            texte = r.json()["content"][0]["text"].strip().strip("```json").strip("```").strip()
            data = json.loads(texte)
            return data.get("faux_positif", True), data.get("raison", "")
        except Exception as e:
            if tentative == 2:
                return True, f"Erreur: {str(e)[:40]}"
            time.sleep(5)
    return True, "Echec apres 3 tentatives"

def enrichir_vrai_positif(nom, siren):
    """Appel Claude AVEC web search — uniquement pour les vrais positifs."""
    prompt = f"""Recherche des informations sur cette entreprise française :
Nom : "{nom}"
SIREN : {siren}

Réponds UNIQUEMENT en JSON :
{{
  "description_activite": "2-3 phrases sur l'activité",
  "chiffre_affaires": "montant avec année ex: 45M€ (2023) ou null",
  "ebitda": "montant avec année ou null",
  "actionnariat": "famille X, fonds Y, groupe coté Z... ou null",
  "derniere_actualite": "dernière actu pertinente (cession, acquisition, levée de fonds, croissance) avec date ou null"
}}"""

    payload = {
        "model": "claude-haiku-4-5-20251001",
        "max_tokens": 1000,
        "tools": [{"type": "web_search_20250305", "name": "web_search"}],
        "messages": [{"role": "user", "content": prompt}],
    }
    for tentative in range(3):
        try:
            r = requests.post(
                "https://api.anthropic.com/v1/messages",
                headers={"x-api-key": ANTHROPIC_KEY, "anthropic-version": "2023-06-01", "content-type": "application/json"},
                json=payload, timeout=60,
            )
            if r.status_code == 429:
                print(f"  Anthropic 429, attente 30s... (tentative {tentative+1})")
                time.sleep(30)
                continue
            r.raise_for_status()
            texte = ""
            for bloc in r.json()["content"]:
                if bloc.get("type") == "text":
                    texte += bloc.get("text", "")
            texte = texte.strip().strip("```json").strip("```").strip()
            return json.loads(texte)
        except Exception as e:
            if tentative == 2:
                return {"description_activite": None, "chiffre_affaires": None,
                        "ebitda": None, "actionnariat": None, "derniere_actualite": None}
            time.sleep(5)
    return {"description_activite": None, "chiffre_affaires": None,
            "ebitda": None, "actionnariat": None, "derniere_actualite": None}

def creer_excel(resultats):
    wb = Workbook()
    ws = wb.active
    ws.title = "Entreprises enrichies"

    HEADER_BG = "1F3864"
    thin = Side(style="thin", color="C0C0C0")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)

    headers = [
        "Nom recherché", "SIREN", "Nom officiel", "Ville", "Faux positif",
        "Raison fiabilité", "Description activité", "CA", "EBITDA",
        "Actionnariat", "Dernière actualité"
    ]
    largeurs = [30, 12, 30, 20, 12, 35, 50, 18, 18, 35, 50]

    for col, (h, w) in enumerate(zip(headers, largeurs), 1):
        cell = ws.cell(row=1, column=col, value=h)
        cell.font = Font(bold=True, color="FFFFFF", name="Arial", size=10)
        cell.fill = PatternFill("solid", start_color=HEADER_BG)
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = border
        ws.column_dimensions[get_column_letter(col)].width = w
    ws.row_dimensions[1].height = 30

    for i, res in enumerate(resultats, 2):
        bg = "EEF2F7" if i % 2 == 0 else "FFFFFF"
        faux = res.get("faux_positif")
        row_data = [
            res["nom_recherche"],
            res.get("siren", ""),
            res.get("nom_trouve", ""),
            res.get("ville", ""),
            "OUI" if faux else ("NON" if faux is False else res.get("statut", "?")),
            res.get("raison", ""),
            res.get("description_activite", ""),
            res.get("chiffre_affaires", ""),
            res.get("ebitda", ""),
            res.get("actionnariat", ""),
            res.get("derniere_actualite", ""),
        ]
        for col, val in enumerate(row_data, 1):
            cell = ws.cell(row=i, column=col, value=val)
            cell.font = Font(name="Arial", size=9)
            cell.fill = PatternFill("solid", start_color=bg)
            cell.alignment = Alignment(vertical="top", wrap_text=True)
            cell.border = border
            if col == 5:
                if val == "OUI":
                    cell.font = Font(name="Arial", size=9, color="CC0000", bold=True)
                elif val == "NON":
                    cell.font = Font(name="Arial", size=9, color="006600")
        ws.row_dimensions[i].height = 70

    ws.freeze_panes = "A2"
    ws.auto_filter.ref = f"A1:{get_column_letter(len(headers))}1"
    wb.save("siren_enrichi_test.xlsx")
    print("Fichier genere : siren_enrichi_test.xlsx")

if __name__ == "__main__":
    resultats = []
    total = len(ENTREPRISES)

    for i, nom in enumerate(ENTREPRISES, 1):
        print(f"\n[{i}/{total}] {nom}")
        res = {"nom_recherche": nom}

        # Étape 1 : SIREN
        siren_res = chercher_siren(nom)
        res.update(siren_res)
        print(f"  SIREN : {siren_res['statut']} → {siren_res.get('siren', '')} {siren_res.get('nom_trouve', '')}")
        time.sleep(3)

        if siren_res["statut"] != "Trouve":
            resultats.append(res)
            continue

        # Étape 2 : Vérification faux positif (sans web search)
        time.sleep(3)  # pause anti-429 Anthropic
        print(f"  Vérification faux positif...")
        faux_positif, raison = verifier_faux_positif(nom, siren_res["nom_trouve"])
        res["faux_positif"] = faux_positif
        res["raison"] = raison
        print(f"  → Faux positif : {faux_positif} — {raison}")

        if faux_positif:
            print(f"  Faux positif détecté, pas d'enrichissement.")
            resultats.append(res)
            continue

        # Étape 3 : Enrichissement web search (seulement vrais positifs)
        print(f"  Enrichissement avec recherche web...")
        enrichi = enrichir_vrai_positif(nom, siren_res["siren"])
        res.update(enrichi)
        print(f"  → CA : {enrichi.get('chiffre_affaires', 'N/A')}")
        print(f"  → Actionnariat : {enrichi.get('actionnariat', 'N/A')}")

        resultats.append(res)

    creer_excel(resultats)
    print(f"\nTest terminé : {total} entreprises traitées.")
