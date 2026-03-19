import requests
import os
import time
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

API_KEY = os.environ["INSEE_KEY"]

ENTREPRISES = [
    "3DEUS DYNAMICS", "A-NSE", "AAS INDUSTRIES", "ABC", "AC DISMANTLING",
    "ACB", "ADB", "ADDEV MATERIALS AEROSPACE SAS", "AddUp", "ADHETEC",
    "ADSS", "AEGIS PLATING", "AEQUS AEROSPACE FRANCE", "AERIADES",
    "AERO NEGOCE INTERNATIONAL", "AEROCAMPUS AQUITAINE", "AEROCAST",
    "AEROCENTRE", "AEROMETAL SAS", "AEROMETALS & ALLOYS",
    "AEROSPACE & DEFENSE OXYGEN SYSTEMS SAS",
    "AEROSPACE CLUSTER AUVERGNE RHONE-ALPES", "AERTEC", "AGENA SPACE",
    "AIF - ATELIERS DE L ILE DE FRANCE", "AIR COST CONTROL", "AIRMOD",
    "AIRUDIT", "AKKODIS AKKA TECHNOLOGIES", "ALLIANCE OUTILLAGE",
    "ALPHA IMPULSION", "ALTYTUD Cluster Aeronautique", "ANTAVIA", "APLAST",
    "APPLUS", "AQL ELECTRONIQUE", "AQUARESE Industries", "Arelis",
    "ARKADIA", "ARMISIA GROUP", "ARTUS", "ASCENDANCE FLIGHT TECHNOLOGIES",
    "ASTROSCALE", "ATELIERS DE LA HAUTE GARONNE", "ATEQ AVIATION",
    "AUXITROL SAS", "AVIATEC", "AVIATUBE", "AVL FRANCE", "AXON CABLE",
    "AXYAL", "BECKER ELECTRONIQUE", "BELINK HIRECT SAS", "BLUE SPIRIT AERO",
    "BOLLHOFF GILLIS", "BRETAGNE AEROSPACE", "BROWN EUROPE", "BT2i Group",
    "BUFAB France", "BUREAU VERITAS", "CAILLAU", "CALIDER",
    "CAPGEMINI ENGINEERING", "CATHERINEAU", "CATOIRE SEMI", "CENTUM T&S",
    "CGI FRANCE", "CICOR COMBREE", "CIR", "CIRCOR INDUSTRIA",
    "CLAYENS GENAS", "COMAT", "COMPTOIR GENERAL DES METAUX",
    "CONSTRUCTION MECANIQUE DE PRECISION", "CS GROUP", "CT INGENIERIE",
    "CTMI", "CURTISS-WRIGHT ARRESTING SYSTEMS SAS", "CYBERMECA", "DAHER",
    "DEFI Group", "DEMGY", "DESHONS HYDRAULIQUE", "DESSIA", "DOMUSA",
    "DRAKA FILECA", "DUQUEINE Atlantique",
    "EATON INTERCONNECT TECHNOLOGIES", "EATON SAS",
    "ECHEVERRIA", "ECM", "EES CLEMESSY", "ELDEC France", "EMITECH",
    "ENNOVI AMS FRANCE", "ESTUAIRE", "ETIM", "ETT", "EUREP Industries",
    "EXAIL AEROSPACE", "EXENS GROUP", "EXOES BATTERY COOLING", "EXPIRIS",
    "EXPLEO", "EXPLISEAT", "EXTENSEE", "EXXELIA", "FAURE HERMAN",
    "FEDERAL MOGUL SYSTEMS PROTECTION", "FERCHAU FRANCE", "FERRY-CAPITAIN",
    "FLEURET", "FLUOR ONE", "FLYING WHALES", "FLYING-ROBOTS HSF",
    "FREGATE", "FREYSSINET AERO EQUIPMENT", "GACHES CHIMIE SPECIALITES",
    "GALILE GROUPE", "GCA Supply PACKING", "GEKATEX GROUP", "GESTAL",
    "GLOBAL BIOENERGIES", "GLOBALSYS", "GMI AERO", "GMP INDUSTRIE",
    "GOODRICH ACTUATION SYSTEMS", "GREENERWAVE", "GRESSET ASSOCIES SAS",
    "GROUPE APAVE", "GROUPE BLONDEL", "GROUPE LPF", "GROUPE ROSSI AERO",
    "Groupe TRA-C industrie", "HALGAND", "HAPSTER",
    "HENKEL TECHNOLOGIES FRANCE", "HEXCEL", "HOWMET FASTENING SYSTEMS",
    "HYBROGINES", "HYNAERO", "HYPRSPACE", "ICM INDUSTRIE", "IDEA LOGISTIQUE",
    "INDRAERO-SIREN", "INFINITY SPACE PROVIDERS",
    "INSTITUT DE SOUDURE INDUSTRIE", "INVENTEC PERFORMANCE CHEMICALS",
    "ION-X", "ISI MIDI-PYRENEES", "JACQUES DUBOIS", "JCM3 SUPERMETAL",
    "JET CUT", "JOGAM", "JONE PRECISION", "JSM PERRIN",
    "KEP TECHNOLOGIES INTEGRATED SYSTEMS", "KEPPLAIR EVOLUTION", "KINEIS",
    "KOMUGI", "KWAN-TEK", "L UNION DES FORGERONS", "LACHANT STAMPING",
    "LAUAK", "LE BOZEC FILTRATION SYSTEMS", "LE CRENEAU INDUSTRIEL",
    "LEFAE", "LEOBLUE", "LHOTELLIER", "LMB", "LOGAERO SERVICES",
    "LOIRETECH INGENIERIE", "LUTRINGER INDUSTRIES", "MADER FRANCE",
    "MANUDEM", "MAP SPACE COATINGS", "MAPAERO", "MASER ENGINEERING",
    "MAXON FRANCE", "MECANIQUE ATELIER DE COIGNIERES", "MECAPOLE",
    "MECAPROTEC Industries", "MEGGITT SENSOREX",
    "MERSEN France Gennevilliers", "METAVONICS", "MICROSTEEL",
    "MILTECH INTERNATIONAL", "MIRATLAS", "MIURA SIMULATION",
    "MK AIR MEKAMICRON", "MONIN MECANIQUE", "MULTIPLAST", "NAE NORMANDIE",
    "NAULUM SOLUTIONS", "NEHIA", "NEOPOLIA AEROSPACE", "NEUROBUS",
    "NEXANS AEROSPACE FRANCE", "NEXESS", "NICOMATIC", "NTN EUROPE",
    "OEMServices SAS", "OERLIKON BALZERS", "OMEGA SYSTEMES ATLANTIQUE",
    "OPLIT", "ORUS", "OTONOMY AVIATION", "OXY SIGN", "OXYTRONIC",
    "PANGEA AEROSPACE FRANCE", "PELICO", "PERCALL", "PGA ELECTRONIC",
    "PINETTE PEI", "PMT ASD", "POCHET AEROSPACE", "PREDELL SERVICES",
    "PRODEX AEROSPACE SOLUTIONS", "PRODUITS PLASTIQUES PERFORMANTS",
    "PROFORM", "PROMETHEE", "RATIER-FIGEAC", "RELLUMIX", "REXIAA", "RIDE",
    "ROCKWELL COLLINS FRANCE", "SACI", "SAMD", "SCA", "SECAMIC",
    "SECRE COMPOSANTS ELECTRONIQUES", "SEGNERE", "SELECTARC GROUP", "SENX",
    "SEREME", "SERMA TECHNOLOGIES", "SFGP", "SIS INDUSTRIE",
    "SKF Aeroengine France", "SKF Aerospace", "SKYREAL", "SMD AERO",
    "SODERN", "SODITECH", "SOGECLAIR AEROSPACE", "SOGITEC Industries",
    "SONOVISION", "SOPHIA ENGINEERING", "SOPRA STERIA Group", "SOREAM",
    "SPACE", "SPACE NETWORK SERVICES", "SPACELOCKER", "SPECITUBES",
    "SPHEREA", "SPIX INDUSTRY", "SREBOT TECHNOLOGIES", "ST GROUP", "STACEM",
    "STARBURST ACCELERATOR", "STEG", "STI FRANCE", "STRATOFLIGHT", "SUNAERO",
    "SUPER BIRDIE", "SURFEO", "T3S TECNIC SERIGRAPHIE SERVICE",
    "TEAM PLASTIQUE", "TECHNI-MODUL ENGINEERING", "TESTIA", "TETMET",
    "THERMI-LOIRE", "TIDAV", "TIKEHAU INVESTMENT MANAGEMENT", "TIMET SAVOIE",
    "TITEFLEX EUROPE", "TRAMEC AERO",
    "TRELLEBORG SEALING SOLUTIONS FRANCE", "TRESCAL",
    "TURGIS ET GAILLARD INDUSTRIE", "TYCO ELECTRONICS FRANCE", "U-Space",
    "UAC CEFIVAL", "ULMER AERONAUTIQUE", "USI", "VIRAJ AERO", "VOLTAERO",
    "WL GORE ASSOCIES", "WALLACE TECHNOLOGIES", "WEISS TECHNIK",
    "WELCO INDUSTRIES", "WHEELABRATOR GROUP", "WINGLEET", "WIREONE INDUSTRY",
    "WORMSENSING", "ZOZIO"
]

HEADERS = {
    "Accept": "application/json",
    "X-INSEE-Api-Key-Integration": API_KEY
}

def chercher_siren(nom):
    nom_clean = nom.strip()
    url = "https://api.insee.fr/api-sirene/3.11/siren"
    params = {
        "q": f'denominationUniteLegale:"{nom_clean}"',
        "nombre": 3,
        "champs": "siren,periodesUniteLegale"
    }
    try:
        r = requests.get(url, headers=HEADERS, params=params, timeout=15)
        if r.status_code == 429:
            print("  Rate limit, attente 60s...")
            time.sleep(60)
            r = requests.get(url, headers=HEADERS, params=params, timeout=15)
        if r.status_code == 200:
            data = r.json()
            unites = data.get("unitesLegales", [])
            if unites:
                u = unites[0]
                periodes = u.get("periodesUniteLegale", [{}])
                denomination = periodes[0].get("denominationUniteLegale", "") if periodes else ""
                etat = periodes[0].get("etatAdministratifUniteLegale", "") if periodes else ""
                return {
                    "siren": u.get("siren", ""),
                    "denomination_officielle": denomination,
                    "etat": "Active" if etat == "A" else "Fermee" if etat == "C" else etat,
                    "nb_resultats": len(unites),
                    "statut": "Trouve"
                }
        elif r.status_code == 404:
            return {"siren": "", "denomination_officielle": "", "etat": "", "nb_resultats": 0, "statut": "Non trouve"}
        else:
            return {"siren": "", "denomination_officielle": "", "etat": "", "nb_resultats": 0, "statut": f"Erreur {r.status_code}"}
    except Exception as e:
        return {"siren": "", "denomination_officielle": "", "etat": "", "nb_resultats": 0, "statut": f"Erreur: {str(e)[:50]}"}

def creer_excel(resultats):
    wb = Workbook()
    ws = wb.active
    ws.title = "SIRENs"

    HEADER_BG = "1F3864"
    thin = Side(style="thin", color="C0C0C0")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)

    headers = ["Nom recherche", "SIREN", "Denomination officielle", "Etat", "Statut", "Nb resultats"]
    largeurs = [45, 15, 45, 12, 15, 15]

    for col, (h, w) in enumerate(zip(headers, largeurs), 1):
        cell = ws.cell(row=1, column=col, value=h)
        cell.font = Font(bold=True, color="FFFFFF", name="Arial", size=10)
        cell.fill = PatternFill("solid", start_color=HEADER_BG)
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.border = border
        ws.column_dimensions[get_column_letter(col)].width = w

    ws.row_dimensions[1].height = 25

    for i, (nom, res) in enumerate(resultats.items(), 2):
        bg = "EEF2F7" if i % 2 == 0 else "FFFFFF"
        row = [nom, res["siren"], res["denomination_officielle"], res["etat"], res["statut"], res["nb_resultats"]]
        for col, val in enumerate(row, 1):
            cell = ws.cell(row=i, column=col, value=val)
            cell.font = Font(name="Arial", size=9)
            cell.fill = PatternFill("solid", start_color=bg)
            cell.alignment = Alignment(vertical="center")
            cell.border = border
            if col == 5 and val == "Non trouve":
                cell.font = Font(name="Arial", size=9, color="CC0000")
            if col == 5 and val == "Trouve":
                cell.font = Font(name="Arial", size=9, color="006600")
        ws.row_dimensions[i].height = 18

    ws.freeze_panes = "A2"
    ws.auto_filter.ref = "A1:F1"

    path = "sirens_entreprises.xlsx"
    wb.save(path)
    print(f"Fichier genere : {path}")
    return path

if __name__ == "__main__":
    resultats = {}
    total = len(ENTREPRISES)
    trouves = 0

    for i, nom in enumerate(ENTREPRISES, 1):
        print(f"[{i}/{total}] {nom}...")
        res = chercher_siren(nom)
        resultats[nom] = res
        if res["statut"] == "Trouve":
            trouves += 1
            print(f"  OK {res['siren']} - {res['denomination_officielle']}")
        else:
            print(f"  KO {res['statut']}")
        time.sleep(2)

    print(f"\n{trouves}/{total} entreprises trouvees")
    creer_excel(resultats)
