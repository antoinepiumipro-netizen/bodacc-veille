import requests
import os
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

API_KEY = os.environ["PAPPERS_KEY"]

# ── CONFIG ──────────────────────────────────────────────
CA_MIN        = 10_000_000   # 10M€ minimum
DEPARTEMENTS  = ["75", "77", "78", "91", "92", "93", "94", "95"]  # Île-de-France
NB_ENTREPRISES = 10          # Passer à 100 ou plus en production
# ────────────────────────────────────────────────────────

def fetch_entreprises():
    resultats = []
    for dept in DEPARTEMENTS:
        params = {
            "api_token": API_KEY,
            "par_page": max(1, NB_ENTREPRISES // len(DEPARTEMENTS)),
            "page": 1,
            "departement": dept,
            "chiffre_affaires_min": CA_MIN,
            "entreprise_cessee": "false",
        }
        try:
            r = requests.get("https://api.pappers.fr/v2/recherche", params=params, timeout=30)
            r.raise_for_status()
            data = r.json()
            entreprises = data.get("resultats", [])
            print(f"  Dept {dept} : {len(entreprises)} entreprises")
            resultats.extend(entreprises)
        except Exception as e:
            print(f"  Erreur dept {dept} : {e}")
    return resultats[:NB_ENTREPRISES]

def get_dirigeant(e):
    dirigeants = e.get("dirigeants", [])
    if dirigeants:
        d = dirigeants[0]
        nom = f"{d.get('prenom', '')} {d.get('nom', '')}".strip()
        return nom
    return ""

def get_actionnariat(e):
    actionnaires = e.get("actionnaires", [])
    if not actionnaires:
        return ""
    parts = []
    for a in actionnaires[:3]:
        nom = a.get("nom_entreprise") or f"{a.get('prenom','')} {a.get('nom','')}".strip()
        pct = a.get("pourcentage_parts", "")
        if pct:
            parts.append(f"{nom} ({pct}%)")
        else:
            parts.append(nom)
    return " / ".join(parts)

def is_filiale_groupe_cote(e):
    actionnaires = e.get("actionnaires", [])
    for a in actionnaires:
        # Heuristique : actionnaire avec > 50% et nom connu de grand groupe
        pct = a.get("pourcentage_parts", 0) or 0
        if float(pct) > 50:
            return True
    return False

def creer_excel(entreprises):
    wb = Workbook()
    ws = wb.active
    ws.title = "Prospection M&A"

    # Couleurs
    HEADER_BG  = "1F3864"  # Bleu foncé
    HEADER_FG  = "FFFFFF"
    ROW_ALT    = "EEF2F7"
    BORDER_COLOR = "C0C0C0"

    # En-têtes dans l'ordre de ton tableau
    headers = [
        "Entreprise", "Out Raison", "SIREN", "Site Internet",
        "Description", "Effectif", "Dates Données Fi.",
        "C.A (M€)", "EBE (M€)", "RN (M€)",
        "Département", "Bureau LFG", "Actionnariat",
        "Coordonnées", "Remarques", "Suivi", "BP", "Date"
    ]

    # Largeurs des colonnes
    largeurs = [30, 20, 15, 25, 30, 10, 18, 12, 12, 12, 15, 15, 35, 25, 20, 10, 10, 12]

    # Header
    thin = Side(style="thin", color=BORDER_COLOR)
    border = Border(left=thin, right=thin, top=thin, bottom=thin)

    for col, (h, w) in enumerate(zip(headers, largeurs), 1):
        cell = ws.cell(row=1, column=col, value=h)
        cell.font = Font(bold=True, color=HEADER_FG, name="Arial", size=10)
        cell.fill = PatternFill("solid", start_color=HEADER_BG)
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = border
        ws.column_dimensions[get_column_letter(col)].width = w

    ws.row_dimensions[1].height = 30

    # Données
    for i, e in enumerate(entreprises, 2):
        siege = e.get("siege", {})
        dept  = siege.get("departement", "")
        ca    = e.get("chiffre_affaires")
        rn    = e.get("resultat_net")
        ca_m  = round(ca / 1_000_000, 1) if ca else ""
        rn_m  = round(rn / 1_000_000, 1) if rn else ""

        # Out raison automatique
        out_raison = ""
        if is_filiale_groupe_cote(e):
            out_raison = "Filiale groupe"

        row = [
            e.get("nom_entreprise", ""),
            out_raison,
            e.get("siren", ""),
            e.get("site_web", ""),
            e.get("forme_juridique", ""),
            e.get("effectif", ""),
            e.get("derniere_mise_a_jour_financiere", ""),
            ca_m,
            "",   # EBE non dispo via API recherche
            rn_m,
            dept,
            "",   # Bureau LFG — à remplir manuellement
            get_actionnariat(e),
            get_dirigeant(e),
            "",   # Remarques
            "",   # Suivi
            "",   # BP
            "",   # Date
        ]

        bg = ROW_ALT if i % 2 == 0 else "FFFFFF"
        for col, val in enumerate(row, 1):
            cell = ws.cell(row=i, column=col, value=val)
            cell.font = Font(name="Arial", size=9)
            cell.fill = PatternFill("solid", start_color=bg)
            cell.alignment = Alignment(vertical="center", wrap_text=True)
            cell.border = border

            # Mise en rouge si Out Raison rempli
            if col == 1 and out_raison:
                cell.font = Font(name="Arial", size=9, color="CC0000")

        ws.row_dimensions[i].height = 20

    # Figer la première ligne
    ws.freeze_panes = "A2"

    # Filtre auto
    ws.auto_filter.ref = f"A1:{get_column_letter(len(headers))}1"

    path = "prospection_MA_IDF.xlsx"
    wb.save(path)
    print(f"✅ Fichier généré : {path}")
    return path

if __name__ == "__main__":
    print("Récupération des entreprises via Pappers...")
    entreprises = fetch_entreprises()
    print(f"\n{len(entreprises)} entreprises récupérées")

    if entreprises:
        creer_excel(entreprises)
    else:
        print("Aucune entreprise trouvée.")
