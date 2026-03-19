import requests
import time
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

BASE_URL = "https://recherche-entreprises.api.gouv.fr/search"
DELAI    = 3   # secondes entre chaque requête

ENTREPRISES = [
    "3DEUS DYNAMICS", "A-NSE", "AAS INDUSTRIES", "ABC", "AC DISMANTLING",
    "ACB", "ADB", "ADDEV MATERIALS AEROSPACE SAS", "AddUp", "ADHETEC",
    "ADSS", "AEGIS PLATING", "AEQUS AEROSPACE FRANCE", "AERIADES",
    "AERO NEGOCE INTERNATIONAL", "AEROCAMPUS AQUITAINE", "AEROCAST",
    "AEROCENTRE", "AEROMETAL SAS", "AEROMETALS & ALLOYS",
    "AEROSPACE & DEFENSE OXYGEN SYSTEMS SAS",
    "AEROSPACE CLUSTER AUVERGNE RHONE-ALPES", "AERTEC", "AGENA SPACE",
    "AIF - ATELIERS DE L'ILE DE FRANCE", "AIR COST CONTROL", "AIRMOD",
    "AIRUDIT", "AKKODIS AKKA TECHNOLOGIES", "ALLIANCE OUTILLAGE",
    "ALPHA IMPULSION", "ALTYTUD Cluster Aéronautique", "ANTAVIA", "APLAST",
    "APPLUS+", "AQL ELECTRONIQUE", "AQUARESE Industries", "Arelis",
    "ARKADIA", "ARMISIA GROUP", "ARTUS", "ASCENDANCE FLIGHT TECHNOLOGIES",
    "ASTROSCALE", "ATELIERS DE LA HAUTE GARONNE", "ATEQ AVIATION",
    "AUXITROL SAS", "AVIATEC", "AVIATUBE", "AVL FRANCE", "AXON'CABLE",
    "AXYAL", "BECKER ELECTRONIQUE", "BELINK HIRECT SAS", "BLUE SPIRIT AERO",
    "BOLLHOFF GILLIS", "BRETAGNE AEROSPACE", "BROWN EUROPE", "BT2i Group",
    "BUFAB France", "BUREAU VERITAS", "CAILLAU", "CALIDER",
    "CAPGEMINI ENGINEERING", "CATHERINEAU", "CATOIRE SEMI", "CENTUM T&S",
    "CGI FRANCE", "CICOR COMBREE", "CIR", "CIRCOR INDUSTRIA",
    "CLAYENS GENAS", "COMAT", "COMPTOIR GENERAL DES METAUX",
    "CONSTRUCTION MECANIQUE DE PRECISION", "CS GROUP", "CT INGENIERIE",
    "CTMI", "CURTISS-WRIGHT ARRESTING SYSTEMS SAS", "CYBERMECA", "DAHER",
    "DEFI Group", "DEMGY", "DESHONS HYDRAULIQUE", "DESSIA", "DOMUSA",
    "DRAKA FILECA", "DUQUEINE Atlantique",
    "EATON INTERCONNECT TECHNOLOGIES DIVISION (ITD)", "EATON SAS",
    "ECHEVERRIA", "ECM", "EES CLEMESSY", "ELDEC France", "EMITECH",
    "ENNOVI AMS FRANCE", "ESTUAIRE", "ETIM", "ETT", "EUREP Industries",
    "EXAIL AEROSPACE", "EXENS GROUP", "EXOES BATTERY & COOLING", "EXPIRIS",
    "EXPLEO", "EXPLISEAT", "EXTENSEE", "EXXELIA", "FAURE HERMAN",
    "FEDERAL MOGUL SYSTEMS PROTECTION", "FERCHAU FRANCE", "FERRY-CAPITAIN",
    "FLEURET", "FLUOR ONE", "FLYING WHALES", "FLYING-ROBOTS / HSF",
    "FREGATE", "FREYSSINET AERO EQUIPMENT", "GACHES CHIMIE SPECIALITES",
    "GALILE GROUPE", "GCA Supply PACKING", "GEKATEX GROUP", "GESTAL",
    "GLOBAL BIOENERGIES", "GLOBALSYS", "GMI AERO", "GMP INDUSTRIE",
    "GOODRICH ACTUATION SYSTEMS", "GREENERWAVE", "GRESSET & ASSOCIES SAS",
    "GROUPE APAVE", "GROUPE BLONDEL", "GROUPE LPF", "GROUPE ROSSI AERO",
    "Groupe TRA-C industrie", "HALGAND", "HAPSTER",
    "HENKEL TECHNOLOGIES FRANCE", "HEXCEL", "HOWMET FASTENING SYSTEMS",
    "HYBROGINES", "HYNAERO", "HYPRSPACE", "ICM INDUSTRIE", "IDEA LOGISTIQUE",
    "INDRAERO-SIREN", "INFINITY SPACE PROVIDERS",
    "INSTITUT DE SOUDURE INDUSTRIE", "INVENTEC PERFORMANCE CHEMICALS",
    "ION-X", "ISI MIDI-PYRENEES", "JACQUES DUBOIS", "JCM3 SUPERMETAL",
    "JET CUT", "JOGAM", "JONE PRECISION", "JSM PERRIN",
    "KEP TECHNOLOGIES INTEGRATED SYSTEMS", "KEPPLAIR EVOLUTION", "KINEIS",
    "KOMUGI", "KWAN-TEK", "L'UNION DES FORGERONS", "LACHANT STAMPING",
    "LAUAK", "LE BOZEC FILTRATION & SYSTEMS", "LE CRENEAU INDUSTRIEL",
    "LEFAE", "LEOBLUE", "LHOTELLIER", "LMB", "LOGAERO SERVICES",
    "LOIRETECH INGENIERIE", "LUTRINGER INDUSTRIES", "MADER FRANCE",
    "MANUDEM", "MAP SPACE COATINGS", "MAPAERO", "MASER ENGINEERING",
    "MAXON FRANCE", "MECANIQUE ATELIER DE COIGNIERES", "MECAPOLE",
    "MECAPROTEC Industries", "MEGGITT SENSOREX",
    "MERSEN France Gennevilliers", "METAVONICS", "MICROSTEEL",
    "MILTECH INTERNATIONAL", "MIRATLAS", "MIURA SIMULATION",
    "MK AIR MEKAMICRON", "MONIN MECANIQUE", "MULTIPLAST", "NAE NORMANDIE",
    "NAULUM SOLUTIONS", "NEHIA", "NEOPOLIA AEROSPACE", "NEUROBUS",
    "NEXANS AEROSPACE FRANCE", "NEXESS", "NICOMATIC", "NTN EUROPE",
    "OEMServices SAS", "OERLIKON BALZERS", "OMEGA SYSTEMES ATLANTIQUE",
    "OPLIT", "ORUS", "OTONOMY AVIATION", "OXY SIGN", "OXYTRONIC",
    "PANGEA AEROSPACE FRANCE", "PELICO", "PERCALL", "PGA ELECTRONIC",
    "PINETTE P.E.I", "PMT ASD", "POCHET AEROSPACE", "PREDELL SERVICES",
    "PRODEX AEROSPACE SOLUTIONS", "PRODUITS PLASTIQUES PERFORMANTS",
    "PROFORM", "PROMETHEE", "RATIER-FIGEAC", "RELLUMIX", "REXIAA", "RIDE !",
    "ROCKWELL COLLINS FRANCE", "SACI", "SAMD", "SCA", "SECAMIC",
    "SECRE COMPOSANTS ELECTRONIQUES", "SEGNERE", "SELECTARC GROUP", "SENX",
    "SEREME", "SERMA TECHNOLOGIES", "SFGP", "SIS INDUSTRIE",
    "SKF Aeroengine France", "SKF Aerospace", "SKYREAL", "SMD AERO",
    "SODERN", "SODITECH", "SOGECLAIR AEROSPACE", "SOGITEC Industries",
    "SONOVISION", "SOPHIA ENGINEERING", "SOPRA STERIA Group", "SOREAM",
    "SPACE", "SPACE NETWORK SERVICES", "SPACELOCKER", "SPECITUBES",
    "SPHEREA", "SPIX INDUSTRY", "SREBOT TECHNOLOGIES", "ST GROUP", "STACEM",
    "STARBURST ACCELERATOR", "STEG", "STI FRANCE", "STRATOFLIGHT", "SUNAERO",
    "SUPER BIRDIE", "SURFEO", "T3S - TECNIC SERIGRAPHIE SERVICE",
    "TEAM PLASTIQUE", "TECHNI-MODUL ENGINEERING", "TESTIA", "TETMET",
    "THERMI-LOIRE", "TIDAV", "TIKEHAU INVESTMENT MANAGEMENT", "TIMET SAVOIE",
    "TITEFLEX EUROPE S.A.S.", "TRAMEC AERO",
    "TRELLEBORG SEALING SOLUTIONS FRANCE", "TRESCAL",
    "TURGIS ET GAILLARD INDUSTRIE", "TYCO ELECTRONICS FRANCE SAS", "U-Space",
    "UAC CEFIVAL", "ULMER AERONAUTIQUE", "USI+", "VIRAJ AERO", "VOLTAERO",
    "W.L. GORE & ASSOCIES", "WALLACE TECHNOLOGIES", "WEISS TECHNIK",
    "WELCO INDUSTRIES", "WHEELABRATOR GROUP", "WINGLEET", "WIREONE INDUSTRY",
    "WORMSENSING", "ZOZIO",
]

def chercher_siren(nom):
    for tentative in range(3):
        try:
            r = requests.get(
                BASE_URL,
                params={"q": nom, "per_page": 3},
                timeout=30
            )
            if r.status_code == 429:
                print(f"  429 rate limit, attente 60s...")
                time.sleep(60)
                continue
            if r.status_code == 200:
                data = r.json()
                resultats = data.get("results", [])
                nb = len(resultats)
                if nb == 0:
                    return None, "Non trouvé", 0
                e = resultats[0]
                siren = e.get("siren", "")
                nom_officiel = e.get("nom_complet", "") or e.get("nom_raison_sociale", "")
                naf = e.get("activite_principale", "")
                ville = e.get("siege", {}).get("libelle_commune", "") if e.get("siege") else ""
                dept = e.get("siege", {}).get("departement", "") if e.get("siege") else ""
                return {
                    "siren": siren,
                    "nom_officiel": nom_officiel,
                    "naf": naf,
                    "ville": ville,
                    "dept": dept,
                    "nb_resultats": nb,
                }, None, nb
            return None, f"Erreur {r.status_code}", 0
        except Exception as e:
            if tentative < 2:
                time.sleep(10)
                continue
            return None, f"Erreur: {str(e)[:50]}", 0
    return None, "Echec après 3 tentatives", 0

def creer_excel(vrais, a_verifier):
    wb = Workbook()

    BG_VRAI = "1F3864"
    BG_FAUX = "8B0000"
    thin = Side(style="thin", color="C0C0C0")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)

    def onglet(ws, titre, couleur, headers, largeurs, lignes):
        ws.title = titre
        for col, (h, w) in enumerate(zip(headers, largeurs), 1):
            cell = ws.cell(row=1, column=col, value=h)
            cell.font = Font(bold=True, color="FFFFFF", name="Arial", size=10)
            cell.fill = PatternFill("solid", start_color=couleur)
            cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
            cell.border = border
            ws.column_dimensions[get_column_letter(col)].width = w
        ws.row_dimensions[1].height = 30
        for i, row in enumerate(lignes, 2):
            bg = "EEF2F7" if i % 2 == 0 else "FFFFFF"
            for col, val in enumerate(row, 1):
                cell = ws.cell(row=i, column=col, value=val)
                cell.font = Font(name="Arial", size=9)
                cell.fill = PatternFill("solid", start_color=bg)
                cell.alignment = Alignment(vertical="center", wrap_text=True)
                cell.border = border
            ws.row_dimensions[i].height = 20
        ws.freeze_panes = "A2"
        ws.auto_filter.ref = f"A1:{get_column_letter(len(headers))}1"

    # Onglet 1 : Vrais positifs (1 seul résultat Sirene)
    ws1 = wb.active
    onglet(ws1, "Vrais positifs - import Orbis", BG_VRAI,
           ["Nom GIFAS", "SIREN", "Nom officiel Sirene", "Code NAF", "Ville", "Département"],
           [30, 12, 35, 12, 25, 12],
           vrais)

    # Onglet 2 : À vérifier manuellement
    ws2 = wb.create_sheet("A vérifier manuellement")
    onglet(ws2, "A vérifier manuellement", BG_FAUX,
           ["Nom GIFAS", "SIREN trouvé", "Nom officiel Sirene", "Raison", "Nb résultats"],
           [30, 12, 35, 40, 12],
           a_verifier)

    wb.save("sirens_gifas.xlsx")
    print(f"\nFichier généré : sirens_gifas.xlsx")
    print(f"  Onglet 1 - Vrais positifs : {len(vrais)}")
    print(f"  Onglet 2 - À vérifier    : {len(a_verifier)}")

if __name__ == "__main__":
    vrais      = []
    a_verifier = []
    total = len(ENTREPRISES)

    for i, nom in enumerate(ENTREPRISES, 1):
        print(f"[{i}/{total}] {nom}...")
        res, erreur, nb = chercher_siren(nom)

        if erreur:
            a_verifier.append([nom, "", "", erreur, 0])
            print(f"  KO — {erreur}")
        elif nb > 1:
            a_verifier.append([nom, res["siren"], res["nom_officiel"],
                               f"{nb} résultats — ambiguïté", nb])
            print(f"  ⚠️  {nb} résultats → à vérifier")
        else:
            vrais.append([nom, res["siren"], res["nom_officiel"],
                          res["naf"], res["ville"], res["dept"]])
            print(f"  OK — {res['siren']} {res['nom_officiel']}")

        time.sleep(DELAI)

    creer_excel(vrais, a_verifier)
